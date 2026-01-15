import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def export_backtest_to_excel(
        filepath,
        summary_stats: pd.DataFrame,
        equity_df: pd.DataFrame,
        weights_df: pd.DataFrame,
        ohlc_data: dict,
        trades_df: pd.DataFrame,
        frontier_df: pd.DataFrame = None
    ):
    """
    Créé un export Excel professionnel avec :
    - Résumé des métriques
    - Courbe du portefeuille + benchmark (graph Excel)
    - Poids des actifs
    - OHLC pour chaque actif
    - Trade log
    - Frontier Markowitz (optionnel)
    """
    wb = Workbook()

    # ============================================================
    # 📌 ONGLET 1 : SUMMARY
    # ============================================================
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Portfolio", "Benchmark", "Diff"])

    for metric in summary_stats.index:
        port = summary_stats.loc[metric]["Portefeuille"]
        bench = summary_stats.loc[metric]["Benchmark"]
        diff = None if (port is None or bench is None) else port - bench
        ws.append([metric, port, bench, diff])

    # Style
    for cell in ws["A"] + ws[1]:
        cell.font = Font(bold=True)

    # Rouge/vert pour les performances
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = Font(color="008000", bold=True)
                elif cell.value < 0:
                    cell.font = Font(color="B00000", bold=True)

    auto_adjust_column_width(ws)

    # ============================================================
    # 📌 ONGLET 2 : PORTFOLIO EQUITY + GRAPH
    # ============================================================
    ws = wb.create_sheet("Portfolio_Equity")
    df = equity_df.copy()

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    auto_adjust_column_width(ws)

    # Graphique Excel
    chart = LineChart()
    chart.title = "Portfolio vs Benchmark"
    chart.y_axis.title = "Value ($)"

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)

    dates = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.set_categories(dates)

    ws.add_chart(chart, "E2")

    # ============================================================
    # 📌 ONGLET 3 : WEIGHTS
    # ============================================================

    ws = wb.create_sheet("Weights")

    for r in dataframe_to_rows(weights_df, index=True, header=True):
        ws.append(r)

    auto_adjust_column_width(ws)

    # # ============================================================
    # # 📌 ONGLET 4 : TRADES
    # # ============================================================
    # ws = wb.create_sheet("Trades")

    # for r in dataframe_to_rows(trades_df, index=False, header=True):
    #     ws.append(r)

    # auto_adjust_column_width(ws)

    # ============================================================
    # 📌 ONGLET 5 : Frontier (Markowitz)
    # ============================================================
    if frontier_df is not None:
        ws = wb.create_sheet("Frontier")
        for r in dataframe_to_rows(frontier_df, index=False, header=True):
            ws.append(r)
        auto_adjust_column_width(ws)

    # ============================================================
    # 📌 ONGLET 6+ : OHLC par actif
    # ============================================================
    for sym, df_asset in ohlc_data.items():
        ws = wb.create_sheet(sym)
        for r in dataframe_to_rows(df_asset, index=True, header=True):
            ws.append(r)
        auto_adjust_column_width(ws)

    # ============================================================
    # Sauvegarde finale
    # ============================================================
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"✔ Excel exporté : {filepath}")
